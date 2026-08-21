from __future__ import annotations

import io
import json
import tempfile
import unittest
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

import storage
import directions
import company_size
from kontur_excel import KonturExcelError, import_kontur_xlsx, parse_kontur_xlsx


HEADERS = [
    "Номер", "ИКЗ", "Название", "НМЦ", "Обеспечение заявки",
    "Обеспечение контракта", "Аванс", "Валюта закупки", "Дата публикации",
    "Окончание приема заявок", "Проведение отбора", "Этап отбора", "Тип торгов",
    "Ссылка на ЕИС", "Способ отбора", "ЭТП", "СМП, СОНО", "Метка ",
    "Комментарий", "Ответственный", "Регион", "Название", "ИНН", "КПП",
    "Место поставки", "Место поставки из документов",
]


class FakeRevenueClient:
    def __init__(self, revenue=25_000_000_000, year=2025, status="found"):
        self.revenue = revenue
        self.year = year
        self.status = status
        self.calls = []

    def lookup(self, inn, company_name=None):
        self.calls.append(inn)
        return company_size.RevenueCheck(
            inn=inn,
            revenue_rub=self.revenue if self.status == "found" else None,
            report_year=self.year if self.status == "found" else None,
            company_name=company_name,
            source="fns-gir-bo",
            status=self.status,
        )

    def lookup_by_name(self, company_name=None):
        self.calls.append(f"name:{company_name}")
        return company_size.RevenueCheck(
            inn="",
            revenue_rub=self.revenue if self.status == "found" else None,
            report_year=self.year if self.status == "found" else None,
            company_name=company_name,
            source="fns-gir-bo-name",
            status=self.status,
        )


def workbook_bytes(eis_url: str | None = None,
                   trade_type: str = "Коммерческие") -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "ДАР коммерческие"
    ws.append(["Закупка"] + [None] * 19 + ["Заказчик"])
    ws.append(HEADERS)
    ws.append([
        "4567280", "", "Внедрение аналитической платформы", 12_500_000, "", "", "",
        "RUB", datetime(2099, 8, 18, 15, 1), datetime(2099, 8, 24, 17, 30), "",
        "Подача заявок", trade_type, eis_url or "", "Запрос предложений", "B2B", "", "",
        "", "", "77 Москва", "ООО «Тестовая крупная компания»", "7709832989",
        "772501001", "Москва", "",
    ])
    ws["A3"].hyperlink = (
        "https://zakupki.kontur.ru/api/redirect?"
        "url=https%3A%2F%2Fzakupki.kontur.ru%2F4567280_458%3Futm_source%3Dexcel&token=temp"
    )
    ws["P3"].hyperlink = (
        "https://zakupki.kontur.ru/api/redirect?"
        "url=https%3A%2F%2Fwww.b2b-center.ru%2Ftender%2F4567280&token=temp"
    )
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def deadline_workbook() -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.append(["Закупка"] + [None] * 19 + ["Заказчик"])
    ws.append(HEADERS)
    rows = [
        ("1", "Разработка информационной системы", datetime(2026, 8, 21, 10, 0)),
        ("2", "Лицензии JMIX", datetime(2026, 8, 19, 10, 0)),
        ("3", "Разработка интернет-портала", datetime(2026, 8, 22, 13, 0)),
        ("4", "Лицензии Postgres Pro", datetime(2026, 8, 17, 18, 0)),
    ]
    for number, title, deadline in rows:
        ws.append([
            number, "", title, 12_500_000, "", "", "", "RUB",
            datetime(2026, 8, 10, 9, 0), deadline, "", "Подача заявок",
            "Коммерческие", "", "Запрос предложений", "B2B", "", "", "", "",
            "77 Москва", "ООО «Тестовая крупная компания»", "7709832989",
            "772501001", "Москва", "",
        ])
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


class KonturExcelTests(unittest.TestCase):
    def test_parse_maps_fields_and_removes_redirect_token(self):
        parsed = parse_kontur_xlsx(workbook_bytes())
        item = parsed["items"][0]
        self.assertEqual(item["tender_id"], "kontur:4567280_458")
        self.assertEqual(item["customer"], "ООО «Тестовая крупная компания»")
        self.assertEqual(item["price_rub"], 12_500_000)
        self.assertEqual(item["deadline"], "2099-08-24T17:30:00")
        self.assertEqual(item["url"], "https://zakupki.kontur.ru/4567280_458?utm_source=excel")
        self.assertNotIn("token=", item["url"])
        self.assertEqual(item["_details"]["customer_inn"], "7709832989")
        self.assertEqual(item["_details"]["etp_url"], "https://www.b2b-center.ru/tender/4567280")
        self.assertIsNone(item["contract_security_pct"])
        self.assertIsNone(item["bid_security_pct"])

    def test_import_is_idempotent_and_keeps_existing_status(self):
        with tempfile.TemporaryDirectory() as tmp:
            conn = storage.connect(str(Path(tmp) / "test.db"))
            client = FakeRevenueClient()
            verified = str(Path(tmp) / "verified.csv")
            first = import_kontur_xlsx(
                workbook_bytes(), conn=conn, revenue_client=client,
                verified_registry_path=verified,
            )
            self.assertEqual((first["new"], first["updated"]), (1, 0))
            self.assertEqual(first["new_ids"], ["kontur:4567280_458"])
            self.assertEqual(first["revenue_api_checks"], 1)
            storage.set_status(conn, "kontur:4567280_458", storage.STATUS_IN_PROGRESS)

            second = import_kontur_xlsx(
                workbook_bytes(), conn=conn, revenue_client=client,
                verified_registry_path=verified,
            )
            self.assertEqual((second["new"], second["updated"]), (0, 1))
            self.assertEqual(second["new_ids"], [])
            self.assertEqual(second["revenue_cache_hits"], 1)
            self.assertEqual(client.calls, ["7709832989"])
            tender = storage.get_tender(conn, "kontur:4567280_458")
            self.assertEqual(tender["status"], storage.STATUS_IN_PROGRESS)
            self.assertEqual(tender["source"], "kontur-excel")
            details = json.loads(tender["details"])
            self.assertEqual(details["etp"], "B2B")
            self.assertEqual(details["customer_revenue_rub"], 25_000_000_000)
            self.assertEqual(details["customer_revenue_year"], 2025)
            conn.close()

    def test_revenue_filter_rejects_small_company(self):
        with tempfile.TemporaryDirectory() as tmp:
            conn = storage.connect(str(Path(tmp) / "test.db"))
            summary = import_kontur_xlsx(
                workbook_bytes(), conn=conn,
                revenue_client=FakeRevenueClient(revenue=9_999_999_999),
                verified_registry_path=str(Path(tmp) / "verified.csv"),
            )
            self.assertEqual(summary["kept"], 0)
            self.assertEqual(summary["skipped_small_company"], 1)
            self.assertEqual(storage.count_all(conn), 0)
            conn.close()

    def test_government_tender_bypasses_company_check_and_has_no_revenue_reason(self):
        class MustNotCallRevenueClient:
            def lookup(self, inn, company_name=None):
                raise AssertionError("государственный тендер не должен проверяться по обороту")

            def lookup_by_name(self, company_name=None):
                raise AssertionError("государственный тендер не должен проверяться по названию")

        with tempfile.TemporaryDirectory() as tmp:
            conn = storage.connect(str(Path(tmp) / "test.db"))
            summary = import_kontur_xlsx(
                workbook_bytes(trade_type="44-ФЗ"),
                conn=conn,
                revenue_client=MustNotCallRevenueClient(),
                verified_registry_path=str(Path(tmp) / "verified.csv"),
            )
            self.assertEqual(summary["kept"], 1)
            self.assertEqual(summary["kept_government"], 1)
            tender = storage.get_tender(conn, "kontur:4567280_458")
            self.assertFalse(any("оборот" in reason.lower()
                                 for reason in tender["reasons"]))
            details = json.loads(tender["details"])
            self.assertEqual(details["customer_turnover_status"],
                             "not_applicable_government")
            conn.close()

    def test_unverified_company_is_removed(self):
        with tempfile.TemporaryDirectory() as tmp:
            conn = storage.connect(str(Path(tmp) / "test.db"))
            summary = import_kontur_xlsx(
                workbook_bytes(), conn=conn,
                revenue_client=FakeRevenueClient(status="no_data"),
                name_registry_path=str(Path(tmp) / "empty.txt"),
                verified_registry_path=str(Path(tmp) / "verified.csv"),
            )
            self.assertEqual(summary["kept"], 0)
            self.assertEqual(summary["kept_unverified_company"], 0)
            self.assertEqual(summary["skipped_unverified_company"], 1)
            self.assertIsNone(storage.get_tender(conn, "kontur:4567280_458"))
            conn.close()

    def test_revenue_service_failure_aborts_without_partial_import(self):
        class FailingClient:
            def lookup(self, inn, company_name=None):
                raise company_size.RevenueServiceError("сервис недоступен")

        with tempfile.TemporaryDirectory() as tmp:
            conn = storage.connect(str(Path(tmp) / "test.db"))
            with self.assertRaises(company_size.RevenueServiceError):
                import_kontur_xlsx(
                    workbook_bytes(), conn=conn, revenue_client=FailingClient(),
                    verified_registry_path=str(Path(tmp) / "verified.csv"),
                )
            self.assertEqual(storage.count_all(conn), 0)
            conn.close()

    def test_rejects_unrelated_workbook(self):
        wb = Workbook()
        wb.active.append(["Не", "Контур"])
        data = io.BytesIO()
        wb.save(data)
        data.seek(0)
        with self.assertRaises(KonturExcelError):
            parse_kontur_xlsx(data)

    def test_deadline_policy_keeps_only_licenses_at_three_days_or_less(self):
        with tempfile.TemporaryDirectory() as tmp:
            conn = storage.connect(str(Path(tmp) / "test.db"))
            client = FakeRevenueClient()
            initial = import_kontur_xlsx(
                deadline_workbook(), conn=conn, now=datetime(2026, 8, 10, 12, 0),
                revenue_client=client,
                verified_registry_path=str(Path(tmp) / "verified.csv"),
            )
            self.assertEqual(initial["kept"], 4)

            short_id = conn.execute(
                "SELECT tender_id FROM tenders WHERE title = ?",
                ("Разработка информационной системы",),
            ).fetchone()[0]
            conn.execute("CREATE TABLE user_favorites (user_id INTEGER, tender_id TEXT)")
            conn.execute("INSERT INTO user_favorites VALUES (1, ?)", (short_id,))
            conn.commit()

            current = import_kontur_xlsx(
                deadline_workbook(), conn=conn, now=datetime(2026, 8, 18, 12, 0),
                revenue_client=client,
                verified_registry_path=str(Path(tmp) / "verified.csv"),
            )
            self.assertEqual(current["kept"], 2)
            self.assertEqual(current["skipped_short_non_license"], 1)
            self.assertEqual(current["skipped_expired"], 1)
            self.assertEqual(current["removed_by_deadline"], 2)
            titles = {row[0] for row in conn.execute("SELECT title FROM tenders")}
            self.assertEqual(titles, {"Лицензии JMIX", "Разработка интернет-портала"})
            self.assertEqual(conn.execute("SELECT COUNT(*) FROM user_favorites").fetchone()[0], 0)
            conn.close()


class DirectionTests(unittest.TestCase):
    def test_context_avoids_known_false_positives(self):
        unrelated = [
            "Захват магнитный PML-300",
            "Шкаф для раздевалок ML-11-30",
            "Термокружка Palermo 480 ml",
            "Внесение изменений в лицензию на пользование недрами",
            "Средства индивидуальной защиты для сотрудников",
            "Мониторинг штрафов ГИБДД и платных дорог",
        ]
        for title in unrelated:
            with self.subTest(title=title):
                self.assertEqual(directions.classify({"title": title}), "other")

    def test_context_recognizes_profile_requests(self):
        expected = {
            "Поставка лицензий Postgres Pro": "license",
            "Разработка цифровой модели технологического процесса": "software",
            "Разработка системы консолидации отчетности": "bi",
            "AI Development Platform для автоматизации разработки": "ml",
            "Разработка информационной системы управления": "software",
            "Мониторинг цен на ТП ML-РЕШЕНИЯ": "ml",
            "Развитие функционала системы Заказчика 1С:УХ": "software",
            "Сертификат технической поддержки СУБД Postgres Pro": "license",
            "Внедрение лицензируемого программного обеспечения для Big Data": "license",
        }
        for title, direction in expected.items():
            with self.subTest(title=title):
                self.assertEqual(directions.classify({"title": title}), direction)


if __name__ == "__main__":
    unittest.main()
