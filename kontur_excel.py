# -*- coding: utf-8 -*-
"""Импорт выгрузок Контур.Закупок в общую модель тендеров приложения."""

from __future__ import annotations

import hashlib
import json
import re
from collections.abc import Callable
from zipfile import BadZipFile
from datetime import date, datetime
from pathlib import PurePosixPath
from urllib.parse import parse_qs, unquote, urlparse

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

import company_size
import directions
import priority_companies
import storage
from icp_config import load_icp
from scoring import score_tender, score_tender_llm


SOURCE_NAME = "kontur-excel"
MAX_ROWS = 100_000
MAX_COLUMNS = 100


class KonturExcelError(ValueError):
    """Понятная пользователю ошибка структуры Excel-выгрузки."""


def _text(value) -> str:
    return "" if value is None else str(value).strip()


def _iso(value) -> str | None:
    if isinstance(value, datetime):
        return value.isoformat(timespec="seconds")
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time()).isoformat(timespec="seconds")
    text = _text(value)
    if not text:
        return None
    for fmt in (None, "%d.%m.%Y %H:%M", "%d.%m.%Y"):
        try:
            parsed = datetime.fromisoformat(text) if fmt is None else datetime.strptime(text, fmt)
            return parsed.isoformat(timespec="seconds")
        except ValueError:
            continue
    return None


def _days_left(deadline: str | None, now: datetime | None = None) -> int | None:
    if not deadline:
        return None
    try:
        return (datetime.fromisoformat(deadline) - (now or datetime.now())).days
    except (TypeError, ValueError):
        return None


def _deadline_disposition(tender: dict, now: datetime) -> str | None:
    """Причина удаления по сроку или None, если тендер нужно оставить."""
    deadline = tender.get("deadline")
    if not deadline:
        return None
    try:
        remaining = datetime.fromisoformat(deadline) - now
    except (TypeError, ValueError):
        return None
    if remaining.total_seconds() < 0:
        return "expired"
    if remaining.days <= 3 and directions.classify(tender) != "license":
        return "short_non_license"
    return None


def _revenue_evidence(check: company_size.RevenueCheck) -> tuple[str, str]:
    if check.source == "local-name-registry":
        return (
            "Компания входит в локальный реестр организаций с оборотом от 10 млрд ₽",
            "крупная компания по реестру",
        )
    year_note = f" за {check.report_year} год" if check.report_year else ""
    return (
        "Размер компании-заказчика больше 10 млрд ₽: "
        f"подтверждён оборот {company_size.revenue_human_value(check.revenue_rub)}"
        f"{year_note}",
        "оборот подтверждён",
    )


def _is_government(details: dict | None) -> bool:
    details = details or {}
    trade_type = str(details.get("trade_type") or "").lower().replace(" ", "")
    government_law = any(marker in trade_type for marker in (
        "44-фз", "223-фз", "615ппрф", "615-пп",
    ))
    return bool(details.get("eis_number") or details.get("eis_url") or government_law)


def _without_revenue_evidence(reasons, labels) -> tuple[list[str], list[str]]:
    clean_reasons = [str(value) for value in (reasons or []) if not (
        str(value).lower().startswith("оборот заказчика") or
        str(value).lower().startswith("размер компании-заказчика") or
        "не удалось подтвердить" in str(value).lower() or
        str(value).lower().startswith("компания входит в локальный реестр") or
        str(value).lower().startswith("заказчик входит в список приоритетных")
    )]
    clean_labels = [str(value) for value in (labels or []) if str(value) not in {
        "оборот подтверждён", "оборот не подтверждён",
        "крупная компания по реестру",
    }]
    return clean_reasons, clean_labels


def _number(value) -> int | float | None:
    if isinstance(value, bool) or value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return value
    cleaned = re.sub(r"[^\d,.-]", "", str(value)).replace(",", ".")
    try:
        result = float(cleaned)
    except ValueError:
        return None
    return int(result) if result.is_integer() else result


def _direct_url(link: str | None) -> str | None:
    """Достаёт постоянный URL из kontur redirect-ссылки с временным token."""
    link = _text(link)
    if not link:
        return None
    try:
        parsed = urlparse(link)
        target = parse_qs(parsed.query).get("url", [None])[0]
        return unquote(target) if target else link
    except (TypeError, ValueError):
        return link


def _hyperlink(cell) -> str | None:
    link = getattr(cell, "hyperlink", None)
    return _direct_url(getattr(link, "target", None)) if link else None


def _stable_id(kontur_url: str | None, number: str, etp: str,
               customer_inn: str, title: str) -> str:
    if kontur_url:
        path = PurePosixPath(urlparse(kontur_url).path)
        slug = path.name.strip()
        if slug:
            return f"kontur:{slug}"[:120]
    raw = "|".join((number, etp, customer_inn, title))
    digest = hashlib.sha256(raw.encode("utf-8")).hexdigest()[:20]
    return f"kontur:{number or digest}:{digest}"[:120]


def _price_display(amount, currency: str) -> str | None:
    if amount is None:
        return None
    suffix = "₽" if currency.upper() in ("", "RUB", "RUR", "РУБ") else currency
    if isinstance(amount, float) and not amount.is_integer():
        shown = f"{amount:,.2f}".replace(",", " ")
    else:
        shown = f"{int(amount):,}".replace(",", " ")
    return f"{shown} {suffix}".strip()


def _header_indexes(ws, header_row: int) -> dict[str, int]:
    headers = [_text(ws.cell(header_row, col).value) for col in range(1, ws.max_column + 1)]

    def one(name: str, occurrence: int = 0, required: bool = False) -> int | None:
        found = [i + 1 for i, value in enumerate(headers) if value == name]
        if len(found) > occurrence:
            return found[occurrence]
        if required:
            raise KonturExcelError(f"В выгрузке не найден обязательный столбец «{name}».")
        return None

    return {
        "number": one("Номер", required=True),
        "title": one("Название", 0, required=True),
        "price": one("НМЦ", required=False),
        "bid_security": one("Обеспечение заявки", required=False),
        "contract_security": one("Обеспечение контракта", required=False),
        "advance": one("Аванс", required=False),
        "currency": one("Валюта закупки", required=False),
        "published": one("Дата публикации", required=True),
        "deadline": one("Окончание приема заявок", required=True),
        "selection_date": one("Проведение отбора", required=False),
        "stage": one("Этап отбора", required=False),
        "trade_type": one("Тип торгов", required=False),
        "eis_url": one("Ссылка на ЕИС", required=False),
        "method": one("Способ отбора", required=False),
        "etp": one("ЭТП", required=False),
        "smp": one("СМП, СОНО", required=False),
        "tag": one("Метка", required=False) or one("Метка ", required=False),
        "comment": one("Комментарий", required=False),
        "responsible": one("Ответственный", required=False),
        "region": one("Регион", required=False),
        "customer": one("Название", 1, required=True),
        "inn": one("ИНН", 0, required=False),
        "kpp": one("КПП", 0, required=False),
        "location": one("Место поставки", required=False),
        "location_docs": one("Место поставки из документов", required=False),
    }


def _find_header_row(ws) -> int:
    for row in range(1, min(ws.max_row, 10) + 1):
        values = {_text(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)}
        if {"Номер", "Окончание приема заявок", "ИНН"}.issubset(values):
            return row
    raise KonturExcelError(
        "Не удалось найти заголовки Контур.Закупок. Нужна исходная выгрузка в формате .xlsx."
    )


def parse_kontur_xlsx(source) -> dict:
    """Читает путь или бинарный поток и возвращает нормализованные записи."""
    try:
        workbook = load_workbook(source, read_only=False, data_only=True)
    except (BadZipFile, InvalidFileException, OSError, ValueError) as exc:
        raise KonturExcelError("Файл не удалось прочитать как Excel (.xlsx).") from exc

    try:
        if not workbook.sheetnames:
            raise KonturExcelError("В Excel-файле нет листов.")
        ws = workbook[workbook.sheetnames[0]]
        if ws.max_row > MAX_ROWS or ws.max_column > MAX_COLUMNS:
            raise KonturExcelError("Выгрузка слишком большая для безопасного импорта.")

        header_row = _find_header_row(ws)
        columns = _header_indexes(ws, header_row)
        records: list[dict] = []
        skipped_blank = 0

        for row_number in range(header_row + 1, ws.max_row + 1):
            def value(key: str):
                col = columns.get(key)
                return ws.cell(row_number, col).value if col else None

            number = _text(value("number"))
            title = _text(value("title"))
            if not number and not title:
                skipped_blank += 1
                continue
            if not number or not title:
                skipped_blank += 1
                continue

            customer = _text(value("customer")) or None
            customer_inn = _text(value("inn"))
            etp = _text(value("etp"))
            number_cell = ws.cell(row_number, columns["number"])
            kontur_url = _hyperlink(number_cell)
            etp_cell = ws.cell(row_number, columns["etp"]) if columns.get("etp") else None
            etp_url = _hyperlink(etp_cell) if etp_cell else None
            eis_value = value("eis_url")
            eis_cell = ws.cell(row_number, columns["eis_url"]) if columns.get("eis_url") else None
            eis_url = ((_hyperlink(eis_cell) if eis_cell else None) or
                       _direct_url(_text(eis_value)) or None)

            amount = _number(value("price"))
            currency = _text(value("currency"))
            rub_amount = amount if currency.upper() in ("", "RUB", "RUR", "РУБ") else None
            published = _iso(value("published"))
            deadline = _iso(value("deadline"))
            location = _text(value("location_docs")) or _text(value("location")) or None
            trade_type = _text(value("trade_type"))
            method = _text(value("method"))

            details = {
                "provider": "Контур.Закупки",
                "kontur_url": kontur_url,
                "etp": etp or None,
                "etp_url": etp_url,
                "customer_inn": customer_inn or None,
                "customer_kpp": _text(value("kpp")) or None,
                "trade_type": trade_type or None,
                "selection_method": method or None,
                "selection_stage": _text(value("stage")) or None,
                "selection_date": _iso(value("selection_date")),
                "smp_sono": _text(value("smp")) or None,
                "eis_url": eis_url,
                "kontur_tag": _text(value("tag")) or None,
                "kontur_comment": _text(value("comment")) or None,
                "kontur_responsible": _text(value("responsible")) or None,
                "bid_security_amount": _number(value("bid_security")),
                "contract_security_amount": _number(value("contract_security")),
                "advance": _number(value("advance")),
            }
            details = {key: val for key, val in details.items() if val is not None}

            records.append({
                "tender_id": _stable_id(kontur_url, number, etp, customer_inn, title),
                "number": number,
                "title": title,
                "url": kontur_url or etp_url,
                "subject": title,
                "customer": customer,
                "region": _text(value("region")) or None,
                "location": location,
                "category": trade_type or method or "Закупка",
                "price_rub": rub_amount,
                "price_display": _price_display(amount, currency),
                "published_at": published,
                "deadline": deadline,
                # В Excel Контур эти поля содержат суммы, а текущая схема ожидает проценты.
                # Суммы сохранены в details, чтобы не показывать их как ошибочные проценты.
                "contract_security_pct": None,
                "bid_security_pct": None,
                "source": SOURCE_NAME,
                "_details": details,
            })

        if not records:
            raise KonturExcelError("В выгрузке не найдено ни одного тендера с номером и названием.")
        return {
            "items": records,
            "sheet": ws.title,
            "rows": ws.max_row - header_row,
            "skipped_blank": skipped_blank,
        }
    finally:
        workbook.close()


def import_kontur_xlsx(source, conn=None, icp: dict | None = None,
                       now: datetime | None = None, use_llm: bool = False,
                       progress_callback: Callable[[dict[str, int]], None] | None = None,
                       revenue_client: company_size.FnsGirBoRevenueClient | None = None,
                       name_registry_path: str | None = None,
                       verified_registry_path: str | None = None) -> dict:
    """Импортирует Excel, применяя текущий фильтр компаний и скоринг приложения."""
    parsed = parse_kontur_xlsx(source)
    icp = icp or load_icp()
    own_connection = conn is None
    conn = conn or storage.connect()
    now = now or datetime.now()
    llm_scorer = None
    if use_llm:
        from LLM_scoring import OpenAITenderScorer

        llm_scorer = OpenAITenderScorer()
    revenue_client = revenue_client or company_size.FnsGirBoRevenueClient()
    kept: list[dict] = []
    skipped_small = 0
    skipped_unverified = 0
    kept_government = 0
    kept_priority = 0
    revenue_api_checks = 0
    revenue_cache_hits = 0
    skipped_expired = 0
    skipped_short = 0
    purge_ids: set[str] = set()
    found = len(parsed["items"])
    processed = 0
    llm_scored = 0
    priority_inns = priority_companies.priority_inns(conn)

    def report_progress() -> None:
        if progress_callback is not None:
            progress_callback({
                "found": found,
                "processed": processed,
                "llm_scored": llm_scored,
                "remaining": found - processed,
            })

    try:
        report_progress()
        # Чистим ранее загруженные из Контура записи по тому же правилу, даже
        # если их уже нет в свежей выгрузке.
        for existing in storage.query_tenders(conn, limit=None):
            if existing.get("source") != SOURCE_NAME:
                continue
            if _deadline_disposition(existing, now):
                purge_ids.add(existing["tender_id"])
            reasons = existing.get("reasons") or []
            labels = existing.get("labels") or []
            details = existing.get("details")
            if isinstance(details, str):
                try:
                    details = json.loads(details)
                except (TypeError, ValueError):
                    details = {}
            status = (details or {}).get("customer_turnover_status")
            is_priority = priority_companies.is_priority_tender(details, priority_inns)
            old_unverified_text = any(
                "нужно подтвердить" in str(reason).lower() or
                "не удалось подтвердить" in str(reason).lower()
                for reason in reasons
            )
            if (not _is_government(details) and not is_priority and
                    (status in {"no_data", "not_found", "invalid_inn", "invalid_name",
                                "name_mismatch"} or
                     "оборот не подтверждён" in labels or old_unverified_text)):
                purge_ids.add(existing["tender_id"])

        for source_item in parsed["items"]:
            item = dict(source_item)
            details = item.pop("_details", {})
            disposition = _deadline_disposition(item, now)
            if disposition == "expired":
                skipped_expired += 1
                purge_ids.add(item["tender_id"])
                processed += 1
                report_progress()
                continue
            if disposition == "short_non_license":
                skipped_short += 1
                purge_ids.add(item["tender_id"])
                processed += 1
                report_progress()
                continue
            is_government = _is_government(details)
            is_priority = priority_companies.is_priority_tender(details, priority_inns)
            revenue_check = None
            if is_government:
                kept_government += 1
            elif is_priority:
                kept_priority += 1
            else:
                revenue_check = company_size.check_revenue_by_inn(
                    conn,
                    details.get("customer_inn"),
                    item.get("customer"),
                    client=revenue_client,
                    now=now,
                    name_registry_path=name_registry_path,
                    verified_registry_path=verified_registry_path,
                )
                if revenue_check.from_cache:
                    revenue_cache_hits += 1
                elif revenue_check.source.startswith("fns-gir-bo"):
                    revenue_api_checks += 1
                if not revenue_check.passes:
                    if revenue_check.is_confirmed:
                        skipped_small += 1
                    else:
                        skipped_unverified += 1
                    purge_ids.add(item["tender_id"])
                    processed += 1
                    report_progress()
                    continue
            purge_ids.discard(item["tender_id"])

            item["days_left"] = _days_left(item.get("deadline"), now)
            result = (
                score_tender_llm(item, icp, scorer=llm_scorer)
                if llm_scorer is not None
                else score_tender(item, icp)
            )
            if llm_scorer is not None:
                llm_scored += 1
            item["score"] = result.score
            item["verdict"] = result.verdict
            if is_government:
                item["reasons"] = list(result.reasons)
                item["labels"] = list(result.labels)
                details.update({
                    "customer_revenue_rub": None,
                    "customer_turnover_rub": None,
                    "customer_revenue_year": None,
                    "customer_turnover_year": None,
                    "customer_revenue_source": "government-exempt",
                    "customer_turnover_source": "government-exempt",
                    "customer_turnover_status": "not_applicable_government",
                    "customer_revenue_checked_at": now.isoformat(timespec="seconds"),
                })
            elif is_priority:
                item["reasons"] = [
                    "Заказчик входит в список приоритетных компаний",
                    *result.reasons,
                ]
                item["labels"] = list(result.labels)
                details.update({
                    "customer_revenue_rub": None,
                    "customer_turnover_rub": None,
                    "customer_revenue_year": None,
                    "customer_turnover_year": None,
                    "customer_revenue_source": "priority-company-exempt",
                    "customer_turnover_source": "priority-company-exempt",
                    "customer_turnover_status": "not_applicable_priority",
                    "customer_revenue_checked_at": now.isoformat(timespec="seconds"),
                })
            else:
                revenue_reason, revenue_label = _revenue_evidence(revenue_check)
                item["reasons"] = [revenue_reason, *result.reasons]
                item["labels"] = [revenue_label, *result.labels]
                details.update({
                    "customer_inn": revenue_check.inn or details.get("customer_inn"),
                    "customer_revenue_rub": revenue_check.revenue_rub,
                    "customer_turnover_rub": revenue_check.revenue_rub,
                    "customer_revenue_year": revenue_check.report_year,
                    "customer_turnover_year": revenue_check.report_year,
                    "customer_revenue_source": revenue_check.source,
                    "customer_turnover_source": revenue_check.source,
                    "customer_turnover_status": revenue_check.status,
                    "customer_revenue_checked_at": now.isoformat(timespec="seconds"),
                })
            item["_details"] = details
            kept.append(item)
            processed += 1
            report_progress()

        removed = storage.delete_tenders(conn, purge_ids, commit=False)
        to_save = [{key: val for key, val in item.items() if key != "_details"} for item in kept]
        saved = storage.save_scored(conn, to_save) if to_save else {"new": [], "updated": []}
        imported_at = now.isoformat(timespec="seconds")
        for item in kept:
            conn.execute(
                "UPDATE tenders SET details = ?, enriched_at = ? WHERE tender_id = ?",
                (json.dumps(item["_details"], ensure_ascii=False), imported_at, item["tender_id"]),
            )
        conn.commit()
        return {
            "sheet": parsed["sheet"],
            "rows": parsed["rows"],
            "parsed": len(parsed["items"]),
            "skipped_blank": parsed["skipped_blank"],
            "skipped_small_company": skipped_small,
            "kept_unverified_company": 0,
            "skipped_unverified_company": skipped_unverified,
            "kept_government": kept_government,
            "kept_priority_company": kept_priority,
            "revenue_api_checks": revenue_api_checks,
            "revenue_cache_hits": revenue_cache_hits,
            "skipped_expired": skipped_expired,
            "skipped_short_non_license": skipped_short,
            "removed_by_deadline": removed,
            "kept": len(kept),
            "new": len(saved.get("new", [])),
            "new_ids": list(saved.get("new", [])),
            "updated": len(saved.get("updated", [])),
        }
    finally:
        if own_connection:
            conn.close()


def recheck_kontur_companies(conn=None,
                             revenue_client: company_size.FnsGirBoRevenueClient | None = None,
                             now: datetime | None = None,
                             name_registry_path: str | None = None,
                             verified_registry_path: str | None = None,
                             progress_callback: Callable[[dict[str, int]], None] | None = None
                             ) -> dict:
    """Повторно проверяет уже сохранённые тендеры и удаляет неподтверждённые."""
    own_connection = conn is None
    conn = conn or storage.connect()
    service = revenue_client or company_size.FnsGirBoRevenueClient()
    now = now or datetime.now()
    rows = [row for row in storage.query_tenders(conn, limit=None)
            if row.get("source") == SOURCE_NAME]
    priority_inns = priority_companies.priority_inns(conn)
    removed_ids: list[str] = []
    kept = 0
    sources: dict[str, int] = {}
    try:
        for index, tender in enumerate(rows, 1):
            details = tender.get("details")
            if isinstance(details, str):
                try:
                    details = json.loads(details)
                except (TypeError, ValueError):
                    details = {}
            details = details or {}
            if _is_government(details):
                reasons, labels = _without_revenue_evidence(
                    tender.get("reasons"), tender.get("labels"),
                )
                details.update({
                    "customer_revenue_rub": None,
                    "customer_turnover_rub": None,
                    "customer_revenue_year": None,
                    "customer_turnover_year": None,
                    "customer_revenue_source": "government-exempt",
                    "customer_turnover_source": "government-exempt",
                    "customer_turnover_status": "not_applicable_government",
                    "customer_revenue_checked_at": now.isoformat(timespec="seconds"),
                })
                conn.execute(
                    "UPDATE tenders SET reasons = ?, labels = ?, details = ? "
                    "WHERE tender_id = ?",
                    (json.dumps(reasons, ensure_ascii=False),
                     json.dumps(labels, ensure_ascii=False),
                     json.dumps(details, ensure_ascii=False), tender["tender_id"]),
                )
                kept += 1
                sources["government-exempt"] = sources.get("government-exempt", 0) + 1
            elif priority_companies.is_priority_tender(details, priority_inns):
                reasons, labels = _without_revenue_evidence(
                    tender.get("reasons"), tender.get("labels"),
                )
                priority_reason = "Заказчик входит в список приоритетных компаний"
                reasons = [priority_reason, *[reason for reason in reasons
                                              if reason != priority_reason]]
                details.update({
                    "customer_revenue_rub": None,
                    "customer_turnover_rub": None,
                    "customer_revenue_year": None,
                    "customer_turnover_year": None,
                    "customer_revenue_source": "priority-company-exempt",
                    "customer_turnover_source": "priority-company-exempt",
                    "customer_turnover_status": "not_applicable_priority",
                    "customer_revenue_checked_at": now.isoformat(timespec="seconds"),
                })
                conn.execute(
                    "UPDATE tenders SET reasons = ?, labels = ?, details = ? "
                    "WHERE tender_id = ?",
                    (json.dumps(reasons, ensure_ascii=False),
                     json.dumps(labels, ensure_ascii=False),
                     json.dumps(details, ensure_ascii=False), tender["tender_id"]),
                )
                kept += 1
                sources["priority-company-exempt"] = (
                    sources.get("priority-company-exempt", 0) + 1
                )
            else:
                check = company_size.check_revenue_by_inn(
                    conn, details.get("customer_inn"), tender.get("customer"),
                    client=service, now=now,
                    name_registry_path=name_registry_path,
                    verified_registry_path=verified_registry_path,
                )
                if not check.passes:
                    removed_ids.append(tender["tender_id"])
                else:
                    reason, label = _revenue_evidence(check)
                    reasons, labels = _without_revenue_evidence(
                        tender.get("reasons"), tender.get("labels"),
                    )
                    details.update({
                        "customer_inn": check.inn or details.get("customer_inn"),
                        "customer_revenue_rub": check.revenue_rub,
                        "customer_turnover_rub": check.revenue_rub,
                        "customer_revenue_year": check.report_year,
                        "customer_turnover_year": check.report_year,
                        "customer_revenue_source": check.source,
                        "customer_turnover_source": check.source,
                        "customer_turnover_status": check.status,
                        "customer_revenue_checked_at": now.isoformat(timespec="seconds"),
                    })
                    conn.execute(
                        "UPDATE tenders SET reasons = ?, labels = ?, details = ? "
                        "WHERE tender_id = ?",
                        (json.dumps([reason, *reasons], ensure_ascii=False),
                         json.dumps([label, *labels], ensure_ascii=False),
                         json.dumps(details, ensure_ascii=False), tender["tender_id"]),
                    )
                    kept += 1
                    sources[check.source] = sources.get(check.source, 0) + 1
            if progress_callback is not None:
                progress_callback({"processed": index, "found": len(rows),
                                   "remaining": len(rows) - index})
        removed = storage.delete_tenders(conn, removed_ids, commit=False)
        conn.commit()
        return {"checked": len(rows), "kept": kept, "removed": removed,
                "sources": sources}
    except Exception:
        conn.rollback()
        raise
    finally:
        if own_connection:
            conn.close()
