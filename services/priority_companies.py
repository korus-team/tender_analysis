# -*- coding: utf-8 -*-
"""Приоритетные компании: точное сопоставление по ИНН и импорт из Excel."""

from __future__ import annotations

import re
from datetime import datetime
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from security_utils import UnsafeArchiveError, validate_zip_archive


MAX_ROWS = 100_000
MAX_COLUMNS = 200


class PriorityCompanyImportError(ValueError):
    """Понятная пользователю ошибка файла со списком компаний."""


def ensure_schema(conn) -> None:
    conn.execute(
        "CREATE TABLE IF NOT EXISTS priority_companies ("
        "id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT NOT NULL, "
        "inn TEXT, created_at TEXT)"
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS priority_companies_inn_idx "
        "ON priority_companies(inn)"
    )
    conn.commit()


def normalize_inn(value) -> str:
    """Нормализует ИНН из текстовой или числовой Excel-ячейки."""
    if value is None or isinstance(value, bool):
        return ""
    if isinstance(value, float):
        if not value.is_integer():
            return ""
        value = int(value)
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    digits = re.sub(r"\D", "", text)
    return digits if len(digits) in (10, 12) else ""


def priority_inns(conn) -> set[str]:
    ensure_schema(conn)
    return {
        inn
        for row in conn.execute("SELECT inn FROM priority_companies WHERE inn IS NOT NULL")
        if (inn := normalize_inn(row["inn"]))
    }


def tender_inn(tender_or_details: dict | None) -> str:
    value = tender_or_details or {}
    details = value.get("details")
    if isinstance(details, dict):
        value = details
    return normalize_inn(value.get("customer_inn"))


def is_priority_tender(tender_or_details: dict | None, inns: set[str]) -> bool:
    inn = tender_inn(tender_or_details)
    return bool(inn and inn in inns)


def _header_key(value) -> str:
    text = str(value or "").strip().lower().replace("ё", "е")
    return re.sub(r"[^a-zа-я0-9]+", "", text)


def _inn_header_score(value) -> int:
    key = _header_key(value)
    if key in {"иннкомпании", "иннорганизации", "иннзаказчика"}:
        return 2
    return 1 if key == "инн" else 0


def _name_header_score(value) -> int:
    key = _header_key(value)
    if key in {
        "названиекомпании", "наименованиекомпании", "названиеорганизации",
        "наименованиеорганизации", "названиезаказчика", "наименованиезаказчика",
    }:
        return 2
    return 1 if key in {
        "название", "наименование", "компания", "организация", "заказчик",
    } else 0


def _find_header(worksheet) -> tuple[int, int, int] | None:
    for row_number, row in enumerate(
            worksheet.iter_rows(min_row=1, max_row=min(50, worksheet.max_row),
                                max_col=min(MAX_COLUMNS, worksheet.max_column),
                                values_only=True), 1):
        inn_candidates = [(_inn_header_score(value), index)
                          for index, value in enumerate(row)
                          if _inn_header_score(value)]
        name_candidates = [(_name_header_score(value), index)
                           for index, value in enumerate(row)
                           if _name_header_score(value)]
        inn_column = max(inn_candidates, default=(0, None))[1]
        name_column = max(name_candidates, default=(0, None))[1]
        if inn_column is not None and name_column is not None and inn_column != name_column:
            return row_number, name_column, inn_column
    return None


def import_xlsx(source, conn) -> dict[str, int | str]:
    """Импортирует компании с любого листа, где найдены колонки названия и ИНН."""
    try:
        validate_zip_archive(
            source,
            max_files=10_000,
            max_uncompressed_bytes=100 * 1024 * 1024,
        )
        workbook = load_workbook(source, read_only=True, data_only=True)
    except (BadZipFile, InvalidFileException, OSError, ValueError, UnsafeArchiveError) as exc:
        raise PriorityCompanyImportError("Файл не удалось прочитать как Excel (.xlsx).") from exc

    try:
        selected = None
        for worksheet in workbook.worksheets:
            if worksheet.max_row > MAX_ROWS or worksheet.max_column > MAX_COLUMNS:
                continue
            header = _find_header(worksheet)
            if header is not None:
                selected = (worksheet, *header)
                break
        if selected is None:
            raise PriorityCompanyImportError(
                "Не найдены колонки «Название» и «ИНН». Названия колонок могут "
                "находиться среди других столбцов файла."
            )

        worksheet, header_row, name_column, inn_column = selected
        existing = priority_inns(conn)
        seen = set(existing)
        added = duplicates = invalid = 0
        now = datetime.now().isoformat(timespec="seconds")
        for row in worksheet.iter_rows(
                min_row=header_row + 1, max_row=worksheet.max_row,
                max_col=max(name_column, inn_column) + 1, values_only=True):
            raw_name = row[name_column] if name_column < len(row) else None
            raw_inn = row[inn_column] if inn_column < len(row) else None
            name = str(raw_name or "").strip()
            inn = normalize_inn(raw_inn)
            if not name and raw_inn in (None, ""):
                continue
            if not name or not inn:
                invalid += 1
                continue
            if inn in seen:
                duplicates += 1
                continue
            conn.execute(
                "INSERT INTO priority_companies (name, inn, created_at) VALUES (?,?,?)",
                (name, inn, now),
            )
            seen.add(inn)
            added += 1
        conn.commit()
        return {
            "sheet": worksheet.title,
            "added": added,
            "duplicates": duplicates,
            "invalid": invalid,
        }
    finally:
        workbook.close()
